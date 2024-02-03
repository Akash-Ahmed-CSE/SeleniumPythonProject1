import openpyxl
class HomePageData:
    test_HomePage_data = [{"Name":"Akash","Email":"akash@gmail.com","Gender":"Male"},
                        {"Name":"Akash2","Email":"akash@gmail.com2","Gender":"Male"},
                        {"Name":"Akash2","Email":"akash@gmail.com2","Gender":"Male"}]

    # @staticmethod
    # def getTestData(test_case_name):
    #     Dict = {}
    #     book = openpyxl.load_workbook("C:\\Users\\Akash\\PycharmProjects\\pythonProject3\\TestData\\PythonDemo.xlsx")
    #     sheet = book.active
    #     for i in range(1, sheet.max_row + 1):
    #         if sheet.cell(row=i, column=1).value == test_case_name:
    #             for j in range(2, sheet.max_column + 1):
    #                 # Dict["Name" = "Akash"]
    #                 Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
    #     print(Dict)