import openpyxl
book = openpyxl.load_workbook("C:\\Users\\Akash\\PycharmProjects\\pythonProject3\\TestData\\PythonDemo.xlsx")
sheet = book.active

#{"Name":"Akash2","Email":"akash@gmail.com2","Gender":"Male"}]
Dict = {}

for i in range(1,sheet.max_row+1):
    if sheet.cell(row = i, column =1).value == "Testcase2":
        for j in range(2,sheet.max_column+1):
            #Dict["Name" = "Akash"]
            Dict[sheet.cell(row=1,column=j).value]=sheet.cell(row=i,column=j).value


